import requests
import json
import openpyxl
from openpyxl.styles import Alignment

#Variaveis
j=3
paises = []
capitais = []
area = []
moeda = []

#Leitura dos dados e armazenamento dos dados em listas
request = requests.get("https://restcountries.com/v3.1/all")
todos = json.loads(request.content)
for i in range (0,len(todos)):
    moeda1 = []
    paises.append(todos[i]['name']['common'])
    try:
      capitais.append(todos[i]['capital'][0])
    except:
      capitais.append("-")
    area.append(todos[i]['area'])
    try:
      dict_moeda = todos[i]['currencies']
      for key in dict_moeda.keys():
        moeda1.append(key)
      moeda.append(moeda1)
    except:
      moeda.append("-")

#Criação inicial do excel
excel = openpyxl.load_workbook("dados.xlsx")
sheet = excel.worksheets[0]
sheet.merge_cells('A1:D1')
sheet.cell(row=1,column=1).value = 'Countries List'
cor = openpyxl.styles.colors.Color(rgb='4F4F4F')
addcor = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=cor)
sheet['A1'].fill = addcor
sheet.cell(row=2,column=1).value = "Name"
sheet.cell(row=2,column=2).value = 'Capital'
sheet.cell(row=2,column=3).value = 'Area'
sheet.cell(row=2,column=4).value = 'Currencies'

cor = openpyxl.styles.colors.Color(rgb='808080')
addcor = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=cor)
sheet['A2'].fill = addcor
sheet['B2'].fill = addcor
sheet['C2'].fill = addcor
sheet['D2'].fill = addcor

currentCell = sheet['A1']
currentCell.alignment = Alignment(horizontal='center')
currentCell = sheet['A2']
currentCell.alignment = Alignment(horizontal='center')
currentCell = sheet['B2']
currentCell.alignment = Alignment(horizontal='center')
currentCell = sheet['C2']
currentCell.alignment = Alignment(horizontal='center')
currentCell = sheet['D2']
currentCell.alignment = Alignment(horizontal='center')

#Adicionando os dados lidos
for i in range (0,len(todos)):
  sheet.cell(row=j,column=1).value = paises[i]
  sheet.cell(row=j,column=2).value = capitais[i]
  sheet.cell(row=j,column=3).value = area[i]
  sheet.cell(row=j,column=4).value = str(moeda[i])
  j = j + 1


excel.save('dados.xlsx')
